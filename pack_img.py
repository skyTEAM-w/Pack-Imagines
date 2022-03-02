import datetime
import os
import re
import shutil
import urllib.error
import urllib.request

import openpyxl
from paddleocr import PaddleOCR

# 全局变量定义
date = datetime.date.today()  # 当日时间
read_path = '.\\list.txt'  # 学生名单存放位置
check_list_path = '.\\{}-{}check_list.txt'.format(date.month, date.day)  # 检查单路径
out_put_path = '{}-{}.txt'.format(date.month, date.day)  # 名单输出路径
main_save_path = '.\\{}{:0>2d}{:0>2d}信息xs2101'.format(date.year, date.month, date.day)  # 主保存路径
save_path = [main_save_path + '\\健康码', main_save_path + '\\行程码', main_save_path + '\\同行密接人员自查']  # 图片保存子路径
check_path = '.\\temp'  # 临时文件保存主路径
temp_path = [check_path + '\\health_code', check_path + '\\tra_code', check_path + '\\close_check']  # 临时文件保存子路径
ID = r'^([1-9]\d{5}[12]\d{3}(0[1-9]|1[012])(0[1-9]|[12][0-9]|3[01])\d{3}[0-9xX])$'  # 身份证匹配字符串
Day = r'{}-{:0>2d}-{:0>2d}.*'.format(date.year, date.month, date.day)  # 健康码 同行密接查询日期匹配
Day2 = r'.*{}\.{:0>2d}\.{:0>2d}.*'.format(date.year, date.month, date.day)  # 行程卡日期匹配
file_name = '{}月{}日信息xs2101“两码一查询”（收集结果）.xlsx'.format(date.month, date.day)  # 收集表关联数据表格名称，与日期有关

# 存储定义
data_dic = {}  # 姓名与[图片下载连接, 临时文件路径]的映射存储
student_list = []  # 名单存储列表
check_list = {'存在问题': [], '未填报': []}  # 检查单两类映射
problem_buf = []  # 检查单文件输入缓存-存在问题
not_submit_buf = []  # 检查单文件输入缓存-上次未填报


# 获取收集结果的函数
def save_data():
    main_book = openpyxl.load_workbook(file_name)
    main_sheet = main_book.active

    for i in range(2, main_sheet.max_row):
        if main_sheet.cell(i, 1).value is not None:
            name = main_sheet.cell(i, 3).value
            url = []

            for j in range(4, main_sheet.max_column):

                if main_sheet.cell(i, j).value is not None:
                    url.append(main_sheet.cell(i, j).hyperlink.target)  # 获取图片链接

            data_dic.update({name: [url,
                                    [str(url[i])[23:66:1].replace('/', '').replace('.', '')
                            .replace('?', '') + '.jpeg'
                                     for i in range(3)]]})  # 定义各图片临时存储位置
            print(i - 1, end=' ')

    print()


def download_img(file_check):
    head = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) "
                      "Chrome/94.0.4606.61 Safari/537.36 Edg/94.0.992.31"}
    for name in data_dic:
        index = 0
        for url in data_dic.get(name)[0]:
            request = urllib.request.Request(url, headers=head)
            try:
                response = urllib.request.urlopen(request)
                img_name = name + '.jpeg'

                if response.getcode() == 200 and ((not os.path.exists(save_path[index] + '\\' + img_name) or name in
                                                   (problem_buf + not_submit_buf))):  # 获取新的图片
                    with open(save_path[index] + '\\' + img_name, 'wb') as file:
                        file.write(response.read())

                shutil.copyfile(save_path[index] + '\\' + img_name,
                                temp_path[index] + '\\' + data_dic.get(name)[1][index])  # 建立临时文件，用于OCR扫描

            except urllib.error.URLError as e:
                if hasattr(e, 'code'):
                    print(e.code)
                if hasattr(e, 'reason'):
                    print(e.reason)
                print('failed!')

            index += 1

        print(name, end=' ')

    # 向检查单写入未填报人员
    for name in student_list:
        if name not in data_dic.keys() and name != '':
            file_check.write(name + '\n')
            check_list['未填报'].append(name)
    file_check.write('-----+-----+-----\n') # 一个分隔符

    # 从未填报上次未填报缓存中去除仍未填报人员
    for name in check_list['未填报']:
        not_submit_buf.pop(not_submit_buf.index(name))

    print()

    print('搞定了')


# 展示一个文字UI
def show_profile():
    print('这是一个临时用于打包打卡数据的python代码')
    print('----------------------------------')
    print('使用前请确保与 .py/.exe 文件同一目录下\n有当日 .xlsx 文件')
    print('一定要使用腾讯文档导出的 .xlsx 文件')
    print('请严格按照QQ群中发送的收集表顺序填写')
    print('遇到bug，及时联系开发人员，比如对面宿舍')
    print('-----------------------------------')
    print('详细阅读后,键入y开始打包,n退出,其他无效')


# 基于PaddleOCR的图片检查方法
def check(file_check, file_out_put):
    print('是否对生成文件进行检查？\ny:是\t其他:否')
    is_stop = False
    problem = {}    # 姓名->问题映射
    while not is_stop:
        choice = str(input())
        if choice == 'y':
            # 若当天第一次进行图片检查，则从数据映射全体扫描，否则根据上次扫描结果进行检查
            for name_t in \
                    (data_dic if len(problem_buf + not_submit_buf) == 0 else (problem_buf + not_submit_buf)):
                error = []  # 错误记录缓存
                index = 0

                # 对健康码与同行密接自查进行检查
                for path in data_dic.get(name_t)[1][::2]:
                    flag1 = False   # 身份证正确标志
                    flag2 = False   # 日期正确标志

                    ocr = PaddleOCR(use_angle_cls=True, lang='ch', use_gpu=True)    # ocr对象
                    result = ocr.ocr(temp_path[index] + '\\' + path, cls=True)      # 结果
                    # 匹配身份证与日期
                    for item in result:
                        if len(re.findall(ID, str(item[1][0]))) > 0:
                            flag1 = True
                        if len(re.findall(Day, str(item[1][0]))) > 0:
                            flag2 = True
                    # 对健康码和同行密接分开处理
                    if index == 0:
                        if not flag1 or not flag2:
                            error.append(save_path[index].replace(main_save_path + '\\', '') + ':' +
                                         ('' if flag1 else '身份证显示不全') + ' ' +
                                         ('' if flag2 else '非当日截图'))
                    elif index == 2:
                        if not flag1:
                            error.append(save_path[index].replace(main_save_path + '\\', '') + ':' +
                                         '身份证显示不全')
                    index += 2

                # 检查行程卡
                ocr = PaddleOCR(use_angle_cls=True, lang='ch', use_gpu=True)
                result = ocr.ocr(temp_path[1] + '\\' + data_dic.get(name_t)[1][1], cls=True)
                flag = False
                for item in result:
                    if len(re.findall(Day2, str(item[1][0]))) > 0:
                        flag = True
                if not flag:
                    error.append(save_path[1].replace(main_save_path + '\\', '') + ':' +
                                 '非当日截图')

                # 若存在问题，向问题字典更新错误
                if len(error) > 0:
                    problem.update({name_t: error})

            is_stop = True
            print('------------------------------------------------------')
            # 向输出文件与检查单文件写入数据
            for name in problem:
                file_check.write(name + '\n')
                file_out_put.write(name + ':')
                print(name + ':', end='')
                for wrong in problem[name]:
                    file_out_put.write(wrong + ' ')
                    print(wrong, end=' ')
                print()
                file_out_put.write('\n')
            print('以上同学可能存在问题，请查看！')
            check_list['存在问题'] = problem
        else:
            is_stop = True


if __name__ == '__main__':
    try:
        is_exit = False
        while not is_exit:
            show_profile()
            select = str(input())
            if select == 'y':

                # 初始化
                # 名单文件初始化
                out_file = open(out_put_path, 'w')
                out_file.write(str(datetime.datetime.now()) + '\n')
                out_file.write('--------------------\n')

                # 检查单初始化
                check_file = None
                if not os.path.exists(check_list_path):
                    check_file = open(check_list_path, 'w+')
                else:
                    check_file = open(check_list_path, 'r')
                    temp_data = check_file.read().split('-----+-----+-----')
                    data_1 = temp_data[0].split('\n')
                    data_2 = temp_data[1].split('\n')
                    for line in data_1:
                        if line != '':
                            not_submit_buf.append(line)
                    for line in data_2:
                        if line != '':
                            problem_buf.append(line)
                # print(problem_buf + not_submit_buf)
                check_file = open(check_list_path, 'w+')    # 更新检查单文件读写状态

                # 读入学生名单
                if os.path.exists(read_path):
                    in_file = open(read_path, 'r')
                    student_list = in_file.read().split('\n')

                # 创建主保存目录与子目录
                if not os.path.exists(main_save_path):
                    os.mkdir(main_save_path)
                    for path in save_path:
                        os.mkdir(path)
                # 创建主临时目录与子目录
                if not os.path.exists(check_path):
                    os.mkdir(check_path)
                    for temp in temp_path:
                        os.mkdir(temp)

                # 实际操作区
                save_data()
                download_img(check_file)
                check(check_file, out_file)

                # 向名单写入未填报名单
                out_file.write('--------------------\n')
                for name in check_list['未填报']:
                    if name != '':
                        out_file.write(name + ':' + '未填报' + '\n')

                is_exit = True

            elif select == 'n':
                is_exit = True

    # 删除临时文件
    finally:
        for root, dirs, files in os.walk(check_path, topdown=False):
            for name in files:
                os.remove(os.path.join(root, name))
            for name in dirs:
                os.rmdir(os.path.join(root, name))
        os.rmdir(check_path)
